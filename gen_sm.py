import openpyxl as op
import json

def gen_sm(y,ws):

    ic=ws.cell(y,3)
    rc=ws.cell(y,4)
    ec=ws.cell(y,5)
    tc=ws.cell(y,6)
    idc=ws.cell(y,7)


    text='{\n    "type": "minecraft:smelting",\n    "ingredient": {\n        "item": "'+ic.value+'"\n    },\n    "result": "'+rc.value+'",\n    "experience": '+str(ec.value)+',\n    "cookingtime": '+str(tc.value)+'\n}'
    text=json.loads(text)

    file_name="recipe/smelting/"+idc.value+".json"
    with open(file_name,'w') as f:
        json.dump(text,f,indent=4)


wb=op.load_workbook("mod.xlsx")
ws=wb['かまど']

y=1
while True:
    y+=1
    x=ws.cell(y,1)
    if(x.value is None):break
    gen_sm(y,ws)
