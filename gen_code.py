import openpyxl as op

wb=op.load_workbook("mod.xlsx")
ws=wb['regi']

def gen_block(y,ws):
    b=ws.cell(y,2)
    c=ws.cell(y,3)
    d=ws.cell(y,4)
    e=ws.cell(y,5)
    f=ws.cell(y,6)
    g=ws.cell(y,7)
    h=ws.cell(y,8)
    i=ws.cell(y,9)

    Bname=str(b.value)
    Sname=str(c.value)
    Material=str(d.value)
    Hard=str(round(e.value,1))
    Burst=str(round(f.value,1))
    Sound=str(g.value)
    Tool=str(h.value)
    Level=str(i.value)

    text='  public static final RegistryObject<Block> '+Bname+'=BLOCKS.register("'+Sname+'",()->new Block(AbstractBlock.Properties.of(Material.'+Material+').requiresCorrectToolForDrops().harvestTool(ToolType.'+Tool+').harvestLevel('+Level+')'

    if(not(Burst is None)):
        text+='.strength('+Hard+'F)'
    else:
        text+='.strength('+Hard+'F,'+Burst+'F)'

    if(Sound is None):
        text+='.sound(SoundType.'+Sound+')'

    text+='));'
    return text

def gen_item(y,ws):
    b=ws.cell(y,2)
    c=ws.cell(y,3)
    d=ws.cell(y,4)
    e=ws.cell(y,5)

    Bname=str(b.value)
    Sname=str(c.value)
    item_type=str(d.value)
    burn=e.value

    text='    public static final RegistryObject<Item> '+Bname+'= ITEMS.register("'+Sname+'",()->new '+item_type+'(Blocks.'+Bname+'.get(),new Item.Properties().tab(CBS_TAB))'

    """
    if(Bname=="Item"):
        text='    public static final RegistryObject<Item> '+Bname+'= ITEMS.register("'+Sname+'",()->new '+item_type+'(new Item.Properties().tab(CBS_TAB))'

    if(not(burn is None)):
        burn=str(round(burn,1))
        text+='{@Override public int getBurnTime(ItemStack itemStack){return '+burn+';}}'
    """
    text+=');'
    return text

def gen_nitem(y,ws):
    b=ws.cell(y,2)
    c=ws.cell(y,3)
    d=ws.cell(y,4)
    e=ws.cell(y,5)

    Bname=str(b.value)
    Sname=str(c.value)
    item_type=str(d.value)
    burn=e.value

    text='    public static final RegistryObject<Item> '+Bname+'= ITEMS.register("'+Sname+'",()->new '+item_type+'(new Item.Properties().tab(CBS_TAB))'
    """
    if(not(burn is None)):
        burn=str(round(burn,1))
        text+='{@Override public int getBurnTime(ItemStack itemStack){return '+burn+';}}'
    """
    text+=');'
    return text

t=0
for i in range(2,60):
    t+=1
    print(gen_item(i,ws))
    print()
    if(t%5==0):
        print("-"*32)

for i in range(60,64):
    t+=1
    print(gen_nitem(i,ws))
    print(t)
    if(t%5==0):
        print("-"*32)

"""
while True:
    y+=1
    ce=ws.cell(y,1)
    if(ce.value is None):break
    print(gen_item(y,ws))
    print()
    if(y>=100):break
"""
