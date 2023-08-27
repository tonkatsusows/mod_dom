import openpyxl as op
import json

def shapeeless(y,ws):
    ac=ws.cell(y,9)
    bc=ws.cell(y,10)
    cc=ws.cell(y,11)
    dc=ws.cell(y,12)
    ec=ws.cell(y,13)
    fc=ws.cell(y,14)
    gc=ws.cell(y,15)
    hc=ws.cell(y,16)
    ic=ws.cell(y,17)

    rec=ws.cell(y,3)
    coc=ws.cell(y,4)
    gro=ws.cell(y,18)
    idc=ws.cell(y,2)

    list=[ac.value,bc.value,cc.value,dc.value,ec.value,fc.value,gc.value,hc.value,ic.value,rec.value,coc.value,gro.value,idc.value,]

    text='{\n    "type": "minecraft:crafting_shapeless",\n    "ingredients": [\n        '

    for i in range(0,9):
        if(not(list[i] is None)):
            text+='\n        {\n            "item": "'+list[i]+'"\n        },'

    text=text[:-1]
    text+='\n    ],\n    "result": {\n        "item": "'+list[9]+'",\n        "count": '+str(list[10])+'\n    }'

    if(not(list[11] is None)):
        text+=',\n    "group": "'+list[11]+'"'

    text+='\n}'

    text=json.loads(text)

    file_name="recipe/shapeeless/"+list[12]+".json"
    with open(file_name,'w') as f:
        json.dump(text,f,indent=4)

def shaped(y,ws):
    ac=ws.cell(y,9)
    bc=ws.cell(y,10)
    cc=ws.cell(y,11)
    dc=ws.cell(y,12)
    ec=ws.cell(y,13)
    fc=ws.cell(y,14)
    gc=ws.cell(y,15)
    hc=ws.cell(y,16)
    ic=ws.cell(y,17)

    rec=ws.cell(y,3)
    coc=ws.cell(y,4)
    gro=ws.cell(y,18)
    idc=ws.cell(y,2)

    r1c=ws.cell(y,6)
    r2c=ws.cell(y,7)
    r3c=ws.cell(y,8)

    list=[ac.value,bc.value,cc.value,dc.value,ec.value,fc.value,gc.value,hc.value,ic.value,rec.value,coc.value,gro.value,idc.value]
    key=["a","b","c","d","e","f","g","h","i"]

    if(not(r1c.value is None)):
        row=[r1c.value]
        l=len(row[0])

    if(not(r2c.value is None)):
        row.append(r2c.value)
        if(len(row[1])>l):l=len(row[1])

    if(not(r3c.value is None)):
        row.append(r3c.value)
        if(len(row[2])>l):l=len(row[2])

    for i in range(0,len(row)):
        if(len(row[i])<l):
            row[i]+=" "

    text='{\n    "type": "minecraft:crafting_shaped",\n    "pattern": [        '

    for x in row:
        text+='\n       "'+x+'",'

    text=text[:-1]

    text+='\n    ],\n    "key":{'

    for i in range(0,9):
        if(not(list[i] is None)):
            text+='\n       "'+key[i]+'": {\n            "item": "'+list[i]+'"\n        },'

    text=text[:-1]
    text+='\n    },\n    "result": {\n        "item": "'+list[9]+'",\n        "count": '+str(list[10])+'\n    }'

    if(not(list[11] is None)):
        text+=',\n    "group": "'+list[11]+'"'

    text+='\n}'

    text=json.loads(text)

    file_name="recipe/shape/"+list[12]+".json"
    with open(file_name,'w') as f:
        json.dump(text,f,indent=4)



wb=op.load_workbook("mod.xlsx")
ws=wb['shape']

y=1

while True:
    y+=1
    c=ws.cell(y,2)
    if(c.value is None):break
    shaped(y,ws)
